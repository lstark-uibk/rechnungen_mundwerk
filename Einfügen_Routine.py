from tkinter import *
from pathlib import Path
import pandas as pd
import openpyxl
import pprint
import dateutil.parser
import datetime
import tkinter as tk
from tkinter import ttk
from tkinter import simpledialog
from PyInquirer import prompt
import pprint
import Einfügen_Routine
import dateutil.parser


def change_place_of_window(root):
    w = 800  # width for the Tk root
    h = 650  # height for the Tk root

    # get screen width and height
    ws = root.winfo_screenwidth()  # width of the screen
    hs = root.winfo_screenheight()  # height of the screen
    # calculate x and y coordinates for the Tk root window
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2)

    # set the dimensions of the screen
    # and where it is placed
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))

def ask_multiple_choice_question(prompt, options):
    root = Tk()
    change_place_of_window(root)
    if prompt:
        Label(root, text=prompt, font=("Helvetica", 14) ).pack()
    v = IntVar()
    for i, option in enumerate(options):
        Radiobutton(root, text=option, variable=v, value=i, font=("Helvetica", 14) ).pack(anchor="w")
    Button(root,text="OK", command=root.destroy,font=("Helvetica", 14) ).pack()
    root.mainloop()
    return options[v.get()]


# das ist alles für den ask many multiple questions
class ScrollFrame(tk.Frame):

    def __init__(self, parent):
        super().__init__(parent)  # create a frame (self)

        self.canvas = tk.Canvas(self, borderwidth=0, background="#ffffff", width=500)  # Canvas to scroll
        self.viewPort = tk.Frame(self.canvas, background="#ffffff")  # This frame will hold the child widgets
        self.vsb = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)  # Attach scrollbar action to scroll of canvas

        self.vsb.pack(side="right", fill="y")  # Pack scrollbar to right - change as needed
        self.canvas.pack(side="left", fill="both",
                         expand=True)  # Pack canvas to left and expand to fill - change as needed
        self.canvas_window = self.canvas.create_window(
            (0, 0),
            window=self.viewPort,
            anchor="nw",
            tags="self.viewPort",
        )  # Add view port frame to canvas

        self.viewPort.bind("<Configure>", self.onFrameConfigure)
        self.canvas.bind("<Configure>", self.onCanvasConfigure)
        self.first = True
        self.onFrameConfigure(None)  # Initial stretch on render

    def onFrameConfigure(self, event):
        '''Reset the scroll region to encompass the inner frame'''
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def onCanvasConfigure(self, event):
        '''Reset the canvas window to encompass inner frame when required'''
        canvas_width = event.width
        self.canvas.itemconfig(self.canvas_window, width=canvas_width)

    def on_mousewheel(self, event):
        '''Allows the mousewheel to control the scrollbar'''
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def bnd_mousewheel(self):
        '''Binds the mousewheel to the scrollbar'''
        self.canvas.bind_all("<MouseWheel>", self.on_mousewheel)

    def unbnd_mousewheel(self):
        '''Unbinds the mousewheel from the scrollbar'''
        self.canvas.unbind_all("<MouseWheel>")

    def delete_all(self):
        '''Removes all widgets from the viewPort, only works if grid was used'''
        children = self.viewPort.winfo_children()
        for child in children:
            child.grid_remove()



def ask_many_multiple_choice_question(prompt, options):
    root = tk.Tk()
    change_place_of_window(root)
    root.title('Frage:')
    if prompt:
        tk.Label(root, text=prompt, font=("Helvetica", 14) ).pack()
    v = tk.IntVar()
    tabs = ttk.Notebook(root)
    tabs.pack()#fill="both")
    list_frame = ttk.Frame(tabs)
    tabs.add(list_frame, text="")

    main_window = list_frame
    mainframe = ttk.Frame(main_window)
    mainframe.grid(column=0, row=0, sticky="W, E, N, S")

    folder_contents_frame = ScrollFrame(mainframe)
    folder_contents_frame.pack(side="top", fill="x", expand=True, padx=20, pady=20)

    for i, option in enumerate(options):
        tk.Radiobutton(folder_contents_frame.viewPort, text=option, variable=v, value=i, font=("Helvetica", 14) ).pack(anchor="w")
    tk.Button(text="OK", command=root.destroy,font=("Helvetica", 14) ).pack()
    root.mainloop()
    return options[v.get()]

def save_to_archive(invoicenumber,datetoday,clientname,invoice_start_date,invoice_end_date,summe,archive_which_invoices_path):
    ws_archive_which_invoices = openpyxl.load_workbook(archive_which_invoices_path)
    archive_which_invoices = ws_archive_which_invoices.worksheets[0]
    invoiceduration = invoice_start_date.strftime("%d.%m.%Y") + " - " + invoice_end_date.strftime("%d.%m.%Y")
    inputdata = [invoicenumber,datetoday, clientname, invoiceduration, summe]

    x = True
    for i in range(1, archive_which_invoices.max_row):
        # define emptiness of cell
        if x == True:
            if archive_which_invoices.cell(i, 1).value is None:
                archive_which_invoices.insert_rows(i, amount=1)
                for col, value in zip(range(1,len(inputdata)+1), inputdata):
                    archive_which_invoices.cell(row=i, column=col, value=value)
                x = False


    ws_archive_which_invoices.save(archive_which_invoices_path)



def get_date():
    import tkinter as tk
    from tkinter import ttk
    from tkcalendar import Calendar, DateEntry

    def cal_done():
        top.withdraw()
        root.quit()

    root = tk.Tk()

    root.withdraw()


    top = tk.Toplevel(root)
    change_place_of_window(top)
    top.attributes("-topmost", True)
    label1 = Label(top, text='Rechnung ab: ', font=("Helvetica", 14) )
    label1.pack(ipadx=10, ipady=10)

    cal1 = Calendar(top,
                   font="Arial 14", selectmode='day')
    cal1.pack(fill="both", expand=True)
    label2 = Label(top, text='bis: ', font=("Helvetica", 14) )
    label2.pack(ipadx=10, ipady=10)
    cal2 = Calendar(top,
                   font="Arial 14", selectmode='day')
    cal2.pack(fill="both", expand=True)

    tk.Button(top, text="ok",height=2, width=20, font="Arial 14", command=cal_done).pack()

    selected_date = None
    root.mainloop()
    return cal1.selection_get(), cal2.selection_get()


def input_new_person(allclientdata_path):
    allclientdata = pd.read_excel(allclientdata_path, index_col=0, header=None, sheet_name=None)

    datatoinquire = list(allclientdata["Vorlage"].index)
    #

    root = Tk()
    # initialise the boxes
    labels = [Label(root, text = onedatalabel) for onedatalabel in datatoinquire]
    entries = [Entry(root) for x in range(0,len(datatoinquire))]


    #position the inquiries in a nice table
    for rownumber, (label, entry) in enumerate(zip(labels, entries)):
        label.grid(column=0, row=rownumber)
        if rownumber != 1 and rownumber != 2:
            entry.grid(column=1, row=rownumber)

    #make the dropdowns
    sexoptions = ["w","m"]
    childoptions =["ja", "nein"]


    child = StringVar(root)
    child.set(childoptions[0])
    childoptiondropdown = OptionMenu(root, child, *childoptions)
    childoptiondropdown.grid(column=1, row=1)

    sex = StringVar(root)
    sex.set(sexoptions[0])
    sexoptiondropdown = OptionMenu(root, sex, *sexoptions)
    sexoptiondropdown.grid(column=1, row=2)


    userinputs = []
    def command():
        for entry in entries:
            userinputs.append(entry.get())
        userinputs[1] = child.get()
        userinputs[2] = sex.get()
        root.destroy()

    Button(root, text="Speichern", command=command).grid(column=1,row =len(datatoinquire)+1)
    root.mainloop()
    # parse datetime inputs
    try:
        userinputs[3] = dateutil.parser.parse(userinputs[3])
    except:
        print("Das Datum ist falsch eingegeben")

    try:
        userinputs[12] = dateutil.parser.parse(userinputs[12])
    finally:

        userinputsdict = dict(zip(datatoinquire, userinputs))


        excelsheet_with_added_person = openpyxl.load_workbook(allclientdata_path)#
        excelsheet_with_added_person.iso_dates = True
        sheet_new_person = excelsheet_with_added_person.create_sheet(userinputsdict["Name"])


        for row, (dataname,userinput) in enumerate(zip(datatoinquire,userinputs)):
            sheet_new_person.cell(row=row+1, column=1).value = dataname
            sheet_new_person.cell(row=row+1, column=2).value = userinput
        excelsheet_with_added_person.save(allclientdata_path)

        print("Ich habe eine neues Blatt für " + userinputsdict["Name"] + " zur PatienInneninformations Exceldatei hinzugefügt")
        print("Mit diesen Einträgen: ")
        pprint.pprint(userinputsdict)
        return userinputsdict


def insert_hourdata(allhourdata_path,clientname):
    root = Tk()
    root.title("Therapiedaten für " + clientname)
    root.geometry("650x500+120+120")

    # empty arrays for your Entrys and StringVars
    text_var = []
    entries = []

    # callback function to get your StringVars
    clienthourdata = []
    def command():
        matrix = []
        for i in range(rows):
            matrix.append([])
            for j in range(cols):
                matrix[i].append(text_var[i][j].get())
        clienthourdata.append(matrix)
        root.destroy()

    labelnames = ["Datum Therapie (im Format wie 1.1.2023)", "Einheitslänge in min"]
    for column in range(0,2):
        Label(root, text=labelnames[column], font=('arial', 10, 'bold'),
          bg="bisque2").place(x=20 + 110*column, y=20)

    x2 = 0
    y2 = 0
    rows, cols = (10,2)
    for i in range(rows):
        # append an empty list to your two arrays
        # so you can append to those later
        text_var.append([])
        entries.append([])
        for j in range(cols):
            # append your StringVar and Entry
            text_var[i].append(StringVar())
            entries[i].append(Entry(root, textvariable=text_var[i][j],width=10))
            entries[i][j].place(x=60 + x2, y=50 + y2)
            x2 += 100

        y2 += 30
        x2 = 0
    button= Button(root,text="Daten speichern", bg='bisque3', width=15, command=command)
    button.place(x=160,y=350)
    root.mainloop()


    clienthourdata = clienthourdata[0]
    datestherapy = list(filter(None,[row[0] for row in clienthourdata] ))
    lengththerapy = list(filter(None,[row[1] for row in clienthourdata]))
    print("unparsed")
    print(datestherapy)
    datestherapy = list(map(lambda x: dateutil.parser.parse(x, dayfirst = True), datestherapy))
    lengththerapy = list(map(lambda x: float(x), lengththerapy))
    print("parsed")
    print(datestherapy)

    excelsheet_hourdata = openpyxl.load_workbook(allhourdata_path)  #
    excelsheet_hourdata.iso_dates = True
    sheet = excelsheet_hourdata["Stundendaten"]

    for dateonetherapy, lengthonetherapy in zip(datestherapy, lengththerapy):
        newRowLocation = sheet.max_row + 1
        sheet.cell(row=newRowLocation, column=1).value = dateonetherapy
        sheet.cell(row=newRowLocation, column=2).value = clientname
        sheet.cell(row=newRowLocation, column=3).value = lengthonetherapy


    excelsheet_hourdata.save(allhourdata_path)
    namehourdata = pd.DataFrame([datestherapy,[clientname for x in range(0,len(datestherapy))],lengththerapy])
    namehourdata = namehourdata.transpose()
    namehourdata.columns = ['Datum', 'Name', 'Minuten']

    return(namehourdata)